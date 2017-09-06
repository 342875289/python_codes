from openpyxl import load_workbook
wb = load_workbook(filename = 'read.xlsx')
sheet_ranges = wb['宣讲会教室借用']
#print(sheet_ranges['A2'].value)

#读第2行第4列单元格
c = sheet_ranges.cell(row=2,column=4)
fill = c.fill
front = c.font
print(c.value)#单元格内容
print(fill.start_color.rgb)#单元格填充颜色
print(front.color.rgb)#单元格字体颜色


c = sheet_ranges.cell(row=7,column=4)
fill = c.fill
front = c.font
print(c.value)
print(fill.start_color.rgb)
print(front.color.rgb)

'''
c = sheet_ranges['E7']
fill = c.fill
front = c.font
print(c.value)
print(fill.start_color.rgb)
print(front.color.rgb)

c = sheet_ranges['F7']
fill = c.fill
front = c.font
print(c.value)
print(fill.start_color.rgb)
print(front.color.rgb)

c = sheet_ranges['G7']
fill = c.fill
front = c.font
print(c.value)
print(fill.start_color.rgb)
print(front.color.rgb)
'''